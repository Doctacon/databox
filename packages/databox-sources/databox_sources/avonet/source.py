"""Pinned AVONET v7 eBird-aligned species-trait source.

Only workbook file 34480856 and worksheet ``AVONET2_eBird`` are accepted.
The signed Figshare storage URL is transient and is never logged or persisted.
"""

from __future__ import annotations

import hashlib
import math
import re
import tempfile
from collections.abc import Callable, Iterator, Mapping
from contextlib import closing
from pathlib import Path
from typing import Any, Protocol, cast
from urllib.error import HTTPError
from urllib.parse import parse_qs, unquote, urlsplit
from urllib.request import HTTPRedirectHandler, Request, build_opener

import dlt
import pendulum
from openpyxl import load_workbook

from databox_sources._logging import get_logger

log = get_logger("databox_sources.avonet")

AVONET_FILE_ID = 34480856
AVONET_SOURCE_URL = "https://ndownloader.figshare.com/files/34480856"
AVONET_EXPECTED_SIZE = 21_524_673
AVONET_EXPECTED_MD5 = "1445afdcfb6df784010c2ca034544bc8"
AVONET_EXPECTED_ROWS = 10_661
AVONET_MAX_BYTES = 24 * 1024 * 1024
AVONET_WORKSHEET = "AVONET2_eBird"
AVONET_DOI = "10.6084/m9.figshare.16586228.v7"
AVONET_VERSION = "v7"
AVONET_LICENSE = "CC BY 4.0"
_CONNECT_TIMEOUT_SECONDS = 10.0
_READ_TIMEOUT_SECONDS = 60.0
_REDIRECT_HOST = "s3-eu-west-1.amazonaws.com"
_REDIRECT_PATH = "/pfigshare-u-files/34480856/AVONETSupplementarydataset1.xlsx"
_AWS_QUERY_KEYS = frozenset(
    {
        "X-Amz-Algorithm",
        "X-Amz-Credential",
        "X-Amz-Date",
        "X-Amz-Expires",
        "X-Amz-SignedHeaders",
        "X-Amz-Signature",
    }
)

EXPECTED_HEADERS = (
    "Species2",
    "Family2",
    "Order2",
    "Avibase.ID2",
    "Total.individuals",
    "Female",
    "Male",
    "Unknown",
    "Complete.measures",
    "Beak.Length_Culmen",
    "Beak.Length_Nares",
    "Beak.Width",
    "Beak.Depth",
    "Tarsus.Length",
    "Wing.Length",
    "Kipps.Distance",
    "Secondary1",
    "Hand-Wing.Index",
    "Tail.Length",
    "Mass",
    "Mass.Source",
    "Mass.Refs.Other",
    "Inference",
    "Traits.inferred",
    "Reference.species",
    "Habitat",
    "Habitat.Density",
    "Migration",
    "Trophic.Level",
    "Trophic.Niche",
    "Primary.Lifestyle",
)

_COLUMNS: Any = {
    "source_scientific_name": {"data_type": "text", "nullable": False},
    "family": {"data_type": "text", "nullable": False},
    "order_name": {"data_type": "text", "nullable": False},
    "avibase_id": {"data_type": "text", "nullable": False, "primary_key": True},
    "total_individuals": {"data_type": "bigint", "nullable": False},
    "female_individuals": {"data_type": "bigint", "nullable": False},
    "male_individuals": {"data_type": "bigint", "nullable": False},
    "unknown_sex_individuals": {"data_type": "bigint", "nullable": False},
    "complete_measures": {"data_type": "bigint", "nullable": False},
    "beak_length_culmen_mm": {"data_type": "double"},
    "beak_length_nares_mm": {"data_type": "double"},
    "beak_width_mm": {"data_type": "double"},
    "beak_depth_mm": {"data_type": "double"},
    "tarsus_length_mm": {"data_type": "double"},
    "wing_length_mm": {"data_type": "double"},
    "kipps_distance_mm": {"data_type": "double"},
    "secondary_length_mm": {"data_type": "double"},
    "hand_wing_index": {"data_type": "double"},
    "tail_length_mm": {"data_type": "double"},
    "mass_g": {"data_type": "double"},
    "mass_source": {"data_type": "text"},
    "mass_reference_other": {"data_type": "text"},
    "inference": {"data_type": "bool", "nullable": False},
    "traits_inferred": {"data_type": "text"},
    "reference_species": {"data_type": "text"},
    "habitat": {"data_type": "text"},
    "habitat_density_code": {"data_type": "bigint", "nullable": False},
    "migration_code": {"data_type": "bigint"},
    "trophic_level": {"data_type": "text"},
    "trophic_niche": {"data_type": "text"},
    "primary_lifestyle": {"data_type": "text", "nullable": False},
    "dataset_doi": {"data_type": "text", "nullable": False},
    "dataset_version": {"data_type": "text", "nullable": False},
    "dataset_license": {"data_type": "text", "nullable": False},
    "source_file_id": {"data_type": "bigint", "nullable": False},
    "source_file_md5": {"data_type": "text", "nullable": False},
    "source_url": {"data_type": "text", "nullable": False},
    "loaded_at": {"data_type": "timestamp", "nullable": False},
}


class BinaryResponse(Protocol):
    status: int
    headers: Mapping[str, str]

    def read(self, __amount: int = -1) -> bytes: ...

    def close(self) -> None: ...


HttpOpen = Callable[[Request, float], BinaryResponse]


class _NoRedirect(HTTPRedirectHandler):
    def redirect_request(
        self, req: Any, fp: Any, code: int, msg: str, headers: Any, newurl: str
    ) -> None:
        return None


def _open_once(request: Request, timeout: float) -> BinaryResponse:
    opener = build_opener(_NoRedirect)
    try:
        return opener.open(request, timeout=timeout)
    except HTTPError as exc:
        if exc.code == 302:
            return cast(BinaryResponse, exc)
        raise


def _validate_redirect_fields(value: str) -> str:
    parsed = urlsplit(value)
    if (
        parsed.scheme != "https"
        or parsed.hostname != _REDIRECT_HOST
        or parsed.username is not None
        or parsed.password is not None
        or parsed.port is not None
        or parsed.fragment
        or unquote(parsed.path) != _REDIRECT_PATH
    ):
        raise ValueError("AVONET download redirect is not approved")
    query = parse_qs(parsed.query, keep_blank_values=True, strict_parsing=True)
    if set(query) != _AWS_QUERY_KEYS or any(len(values) != 1 for values in query.values()):
        raise ValueError("AVONET signed redirect query is not approved")
    algorithm = query["X-Amz-Algorithm"][0]
    credential = query["X-Amz-Credential"][0]
    signed_at = query["X-Amz-Date"][0]
    expires = query["X-Amz-Expires"][0]
    signed_headers = query["X-Amz-SignedHeaders"][0]
    signature = query["X-Amz-Signature"][0]
    credential_match = re.fullmatch(
        r"[A-Z0-9]{16,128}/(\d{8})/eu-west-1/s3/aws4_request", credential
    )
    if (
        algorithm != "AWS4-HMAC-SHA256"
        or signed_headers != "host"
        or credential_match is None
        or re.fullmatch(r"\d{8}T\d{6}Z", signed_at) is None
        or signed_at[:8] != credential_match.group(1)
        or re.fullmatch(r"[0-9]{1,2}", expires) is None
        or not 1 <= int(expires) <= 60
        or re.fullmatch(r"[0-9a-fA-F]{64}", signature) is None
    ):
        raise ValueError("AVONET signed redirect fields are not approved")
    return value


def _validated_redirect(value: str) -> str:
    try:
        return _validate_redirect_fields(value)
    except Exception:
        raise ValueError("AVONET signed redirect is not approved") from None


def download_avonet_workbook(
    destination: Path,
    *,
    http_open: HttpOpen = _open_once,
    expected_size: int = AVONET_EXPECTED_SIZE,
    expected_md5: str = AVONET_EXPECTED_MD5,
) -> None:
    """Download and validate the pinned workbook without retaining its signed URL."""

    initial_request = Request(
        AVONET_SOURCE_URL,
        headers={"Accept": "application/octet-stream", "User-Agent": "databox/1"},
    )
    try:
        initial_response = http_open(initial_request, _CONNECT_TIMEOUT_SECONDS)
    except Exception:
        raise ValueError("AVONET source request failed") from None
    with closing(initial_response) as initial:
        if initial.status != 302:
            raise ValueError("AVONET source did not return the required redirect")
        location = initial.headers.get("Location")
        if not isinstance(location, str):
            raise ValueError("AVONET source redirect is missing")
        target = _validated_redirect(location)

    target_request = Request(
        target,
        headers={"Accept": "application/octet-stream", "User-Agent": "databox/1"},
    )
    digest = hashlib.md5(usedforsecurity=False)
    received = 0
    try:
        target_response = http_open(target_request, _READ_TIMEOUT_SECONDS)
    except Exception:
        raise ValueError("AVONET storage request failed") from None
    try:
        with closing(target_response) as response:
            if response.status != 200 or response.headers.get("Location") is not None:
                raise ValueError("AVONET storage response is not final")
            declared = response.headers.get("Content-Length")
            if declared is not None and (not declared.isdigit() or int(declared) != expected_size):
                raise ValueError("AVONET workbook content length does not match")
            with destination.open("wb") as output:
                while chunk := response.read(64 * 1024):
                    received += len(chunk)
                    if received > AVONET_MAX_BYTES:
                        raise ValueError("AVONET workbook exceeds response limit")
                    digest.update(chunk)
                    output.write(chunk)
    except ValueError:
        raise
    except Exception:
        raise ValueError("AVONET storage response failed") from None
    if received != expected_size:
        raise ValueError("AVONET workbook byte length does not match")
    if digest.hexdigest() != expected_md5:
        raise ValueError("AVONET workbook hash does not match")


def _null(value: object) -> object | None:
    if value == "NA" or (isinstance(value, str) and value.strip() == "") or value is None:
        return None
    return value


def _text(value: object, field: str, *, required: bool = False) -> str | None:
    value = _null(value)
    if value is None:
        if required:
            raise ValueError(f"AVONET {field} is required")
        return None
    if not isinstance(value, str):
        raise ValueError(f"AVONET {field} must be text")
    return value


def _integer(
    value: object, field: str, *, minimum: int, maximum: int, nullable: bool = False
) -> int | None:
    value = _null(value)
    if value is None and nullable:
        return None
    if isinstance(value, bool) or not isinstance(value, int | float) or int(value) != value:
        raise ValueError(f"AVONET {field} must be an integer")
    parsed = int(value)
    if not minimum <= parsed <= maximum:
        raise ValueError(f"AVONET {field} is outside bounds")
    return parsed


def _number(value: object, field: str) -> float | None:
    value = _null(value)
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, int | float):
        raise ValueError(f"AVONET {field} must be numeric")
    parsed = float(value)
    if not math.isfinite(parsed) or not 0 <= parsed <= 10_000_000:
        raise ValueError(f"AVONET {field} is outside bounds")
    return parsed


def _inference(value: object) -> bool:
    if value == "YES":
        return True
    if value == "NO":
        return False
    raise ValueError("AVONET Inference must be YES or NO")


def _normalize_row(values: tuple[object, ...], *, loaded_at: str) -> dict[str, Any]:
    row = dict(zip(EXPECTED_HEADERS, values, strict=True))
    return {
        "source_scientific_name": _text(row["Species2"], "Species2", required=True),
        "family": _text(row["Family2"], "Family2", required=True),
        "order_name": _text(row["Order2"], "Order2", required=True),
        "avibase_id": _text(row["Avibase.ID2"], "Avibase.ID2", required=True),
        "total_individuals": _integer(
            row["Total.individuals"], "Total.individuals", minimum=0, maximum=1_000_000
        ),
        "female_individuals": _integer(row["Female"], "Female", minimum=0, maximum=1_000_000),
        "male_individuals": _integer(row["Male"], "Male", minimum=0, maximum=1_000_000),
        "unknown_sex_individuals": _integer(
            row["Unknown"], "Unknown", minimum=0, maximum=1_000_000
        ),
        "complete_measures": _integer(
            row["Complete.measures"], "Complete.measures", minimum=0, maximum=1_000_000
        ),
        "beak_length_culmen_mm": _number(row["Beak.Length_Culmen"], "Beak.Length_Culmen"),
        "beak_length_nares_mm": _number(row["Beak.Length_Nares"], "Beak.Length_Nares"),
        "beak_width_mm": _number(row["Beak.Width"], "Beak.Width"),
        "beak_depth_mm": _number(row["Beak.Depth"], "Beak.Depth"),
        "tarsus_length_mm": _number(row["Tarsus.Length"], "Tarsus.Length"),
        "wing_length_mm": _number(row["Wing.Length"], "Wing.Length"),
        "kipps_distance_mm": _number(row["Kipps.Distance"], "Kipps.Distance"),
        "secondary_length_mm": _number(row["Secondary1"], "Secondary1"),
        "hand_wing_index": _number(row["Hand-Wing.Index"], "Hand-Wing.Index"),
        "tail_length_mm": _number(row["Tail.Length"], "Tail.Length"),
        "mass_g": _number(row["Mass"], "Mass"),
        "mass_source": _text(row["Mass.Source"], "Mass.Source"),
        "mass_reference_other": _text(row["Mass.Refs.Other"], "Mass.Refs.Other"),
        "inference": _inference(row["Inference"]),
        "traits_inferred": _text(row["Traits.inferred"], "Traits.inferred"),
        "reference_species": _text(row["Reference.species"], "Reference.species"),
        "habitat": _text(row["Habitat"], "Habitat"),
        "habitat_density_code": _integer(
            row["Habitat.Density"], "Habitat.Density", minimum=1, maximum=3
        ),
        "migration_code": _integer(
            row["Migration"], "Migration", minimum=1, maximum=3, nullable=True
        ),
        "trophic_level": _text(row["Trophic.Level"], "Trophic.Level"),
        "trophic_niche": _text(row["Trophic.Niche"], "Trophic.Niche"),
        "primary_lifestyle": _text(row["Primary.Lifestyle"], "Primary.Lifestyle", required=True),
        "dataset_doi": AVONET_DOI,
        "dataset_version": AVONET_VERSION,
        "dataset_license": AVONET_LICENSE,
        "source_file_id": AVONET_FILE_ID,
        "source_file_md5": AVONET_EXPECTED_MD5,
        "source_url": AVONET_SOURCE_URL,
        "loaded_at": loaded_at,
    }


def parse_avonet_workbook(
    path: Path,
    *,
    loaded_at: str,
    expected_rows: int = AVONET_EXPECTED_ROWS,
) -> list[dict[str, Any]]:
    """Parse the one approved sheet after byte-level validation."""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        raise ValueError("AVONET workbook is malformed") from None
    try:
        if workbook.sheetnames.count(AVONET_WORKSHEET) != 1:
            raise ValueError("AVONET worksheet contract does not match")
        worksheet = workbook[AVONET_WORKSHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers != EXPECTED_HEADERS:
            raise ValueError("AVONET header contract does not match")
        normalized: list[dict[str, Any]] = []
        names: set[str] = set()
        identifiers: set[str] = set()
        for values in rows:
            if len(values) != len(EXPECTED_HEADERS):
                raise ValueError("AVONET row width does not match")
            item = _normalize_row(values, loaded_at=loaded_at)
            name = item["source_scientific_name"]
            identifier = item["avibase_id"]
            if name in names or identifier in identifiers:
                raise ValueError("AVONET source names and identifiers must be unique")
            names.add(name)
            identifiers.add(identifier)
            normalized.append(item)
        if len(normalized) != expected_rows:
            raise ValueError("AVONET row count does not match")
        return normalized
    finally:
        workbook.close()


@dlt.resource(
    name="species_traits",
    write_disposition="replace",
    primary_key="avibase_id",
    columns=_COLUMNS,
)
def species_traits() -> Iterator[dict[str, Any]]:
    loaded_at = pendulum.now("UTC").isoformat()
    with tempfile.TemporaryDirectory(prefix="databox-avonet-") as directory:
        workbook_path = Path(directory) / "avonet.xlsx"
        download_avonet_workbook(workbook_path)
        rows = parse_avonet_workbook(
            workbook_path,
            loaded_at=loaded_at,
            expected_rows=AVONET_EXPECTED_ROWS,
        )
        log.info("avonet_workbook_validated", row_count=len(rows), file_id=AVONET_FILE_ID)
        yield from rows


@dlt.source(name="avonet_source")
def avonet_source() -> Any:
    yield species_traits
